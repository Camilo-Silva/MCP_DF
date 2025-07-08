import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import os
from typing import List, Dict, Union
from server import mcp

@mcp.tool()
def exportar_datos_a_excel(
    data: List[Dict] | None = None,
    nombre_archivo: str = "export.xlsx",
    nombre_hoja: str = "Datos",
    incluir_resumen: bool = False,
    columnas_numericas: List[str] | None = None
) -> str:
    """
    Exporta datos (lista de diccionarios) a un archivo Excel con formato profesional.
    
    Args:
        data: Los datos a exportar como lista de diccionarios
        nombre_archivo: Nombre del archivo Excel a crear
        nombre_hoja: Nombre de la hoja dentro del archivo Excel
        incluir_resumen: Si incluir una hoja de resumen automático
        columnas_numericas: Lista de columnas que contienen datos numéricos para formateo
    
    Returns:
        Mensaje indicando éxito o error de la exportación
    """
    try:
        if not data or not isinstance(data, list):
            return "Error: Los datos deben ser una lista de diccionarios no vacía."
        
        if not all(isinstance(item, dict) for item in data):
            return "Error: Todos los elementos deben ser diccionarios."
        
        # Preservar el orden de las columnas del primer diccionario
        if data:
            # Obtener el orden de las columnas del primer diccionario
            column_order = list(data[0].keys())
            # Convertir a DataFrame preservando el orden
            df = pd.DataFrame(data, columns=column_order)
        else:
            df = pd.DataFrame(data)
            column_order = list(df.columns)
        
        if df.empty:
            return "Error: No hay datos para exportar."
        
        # Crear ruta de descarga
        downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
        
        # Asegurar que termine en .xlsx
        if not nombre_archivo.endswith('.xlsx'):
            nombre_archivo += '.xlsx'
        
        # Agregar timestamp para evitar conflictos
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base_name = nombre_archivo.replace('.xlsx', '')
        nombre_archivo_final = f"{base_name}_{timestamp}.xlsx"
        
        ruta_completa = os.path.join(downloads_path, nombre_archivo_final)
        
        # Crear el archivo Excel
        wb = openpyxl.Workbook()
        
        # Eliminar la hoja por defecto
        wb.remove(wb.active)
        
        # === HOJA PRINCIPAL: DATOS ===
        ws_main = wb.create_sheet(nombre_hoja)
        
        # Escribir headers en el orden original preservado
        ws_main.append(column_order)
        
        # Aplicar formato a headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col_num, header in enumerate(column_order, 1):
            cell = ws_main.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Escribir datos en el orden correcto
        for _, row in df.iterrows():
            # Asegurar que los datos se escriban en el orden correcto
            row_data = [row[col] for col in column_order]
            ws_main.append(row_data)
        
        # === HOJA DE RESUMEN (si se solicita) ===
        if incluir_resumen and columnas_numericas:
            ws_resumen = wb.create_sheet("Resumen")
            
            # Headers del resumen
            resumen_headers = ["Métrica", "Valor"]
            ws_resumen.append(resumen_headers)
            
            # Aplicar formato a headers del resumen
            for col_num, header in enumerate(resumen_headers, 1):
                cell = ws_resumen.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Calcular estadísticas
            total_registros = len(df)
            ws_resumen.append(["Total de registros", total_registros])
            
            for col in columnas_numericas:
                if col in df.columns:
                    total = df[col].sum()
                    promedio = df[col].mean()
                    ws_resumen.append([f"Total {col}", round(total, 2)])
                    ws_resumen.append([f"Promedio {col}", round(promedio, 2)])
        
        # Ajustar ancho de columnas para todas las hojas
        for ws in wb.worksheets:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Guardar el archivo
        wb.save(ruta_completa)
        
        # Crear mensaje de éxito
        resultado = "✅ **EXPORTACIÓN EXITOSA A EXCEL**\n\n"
        resultado += f"📄 **Archivo generado:** {nombre_archivo_final}\n"
        resultado += f"📂 **Ubicación:** {ruta_completa}\n\n"
        resultado += f"📊 **Contenido exportado:**\n"
        resultado += f"• Registros totales: {len(df)}\n"
        resultado += f"• Columnas: {len(df.columns)}\n"
        resultado += f"• Hojas: {len(wb.worksheets)}\n"
        resultado += f"• Fecha de exportación: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"
        resultado += "💡 **El archivo se guardó en la carpeta Descargas con formato profesional.**"
        
        return resultado
        
    except Exception as e:
        return f"Error al exportar datos a Excel: {str(e)}"

class ExcelExporterTool:
    """
    Clase utilitaria para exportar datos a Excel de forma genérica.
    Puede ser usada por otras herramientas sin necesidad del decorador @mcp.tool()
    """
    
    @staticmethod
    def export_data(
        data: Union[List[Dict], pd.DataFrame], 
        filename: str = "export.xlsx", 
        sheet_name: str = "Sheet1",
        downloads_folder: bool = True
    ) -> str:
        """
        Exporta datos a un archivo Excel.

        Args:
            data: Los datos a exportar, lista de diccionarios o DataFrame
            filename: Nombre del archivo Excel a crear
            sheet_name: Nombre de la hoja dentro del archivo Excel
            downloads_folder: Si True, guarda en carpeta Descargas del usuario

        Returns:
            String indicando éxito o fallo de la exportación
        """
        try:
            if isinstance(data, list) and all(isinstance(item, dict) for item in data):
                df = pd.DataFrame(data)
            elif isinstance(data, pd.DataFrame):
                df = data
            else:
                return "Error: Los datos deben ser una lista de diccionarios o un DataFrame de pandas."

            if df.empty:
                return "Error: No hay datos para exportar."

            # Determinar ruta de guardado
            if downloads_folder:
                downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
                filepath = os.path.join(downloads_path, filename)
            else:
                filepath = filename

            # Exportar con formato básico
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Aplicar formato básico
                workbook = writer.book
                worksheet = writer.sheets[sheet_name]
                
                # Formatear headers
                for cell in worksheet["1:1"]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")

            return f"Datos exportados exitosamente a {filepath} en la hoja '{sheet_name}'."
            
        except Exception as e:
            return f"Error exportando datos a Excel: {str(e)}"
